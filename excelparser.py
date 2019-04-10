print ("I can finally run")

import openpyxl 
# import Workbook
# wb = Workbook()

# grab the active worksheet
# ws = wb.active

# getting the first required excel file from the folder to compare folders
getFirstFile = openpyxl.load_workbook('C:\Grameenphone\ExcelDataAnalysis\Test1.xlsx')

getSecondFile = openpyxl.load_workbook('C:\Grameenphone\ExcelDataAnalysis\Test2.xlsx')

#the current sheets are active
sheetForFile1 = getFirstFile.active
sheetForFile2 = getSecondFile.active

#compare the first row according to the employee id
# for columns in sheetForFile1
#     getEmployeeID = sheetForFile1.cell(row=1, column=columns)

#     print(getEmployeeID.value)

for row in sheetForFile1.iter_rows():
    for cell in row:
        if cell.value =='EMPLOYEE NO':
            employeeNo_columnNo = cell.column
            employeeNo_rowNo = cell.row
            print (employeeNo_columnNo)
            print (employeeNo_rowNo)
# row+=1




#getting cell data from the file    
# getdata = sheetForFile1.cell(row=1, column=1)

# a = sheetForFile1['A2']
# a3 = sheetForFile1.cell(row=3, column=1)

# print(getdata.value)
# print(a2.value) 
# print(a3.value)

# Data can be assigned directly to cells
# ws['A1'] = 42

# Rows can also be appended
# ws.append([1, 2, 3])

# Python types will automatically be converted
# import datetime
# ws['A2'] = datetime.datetime.now()

# Save the file
# wb.save("Test1.xlsx")

